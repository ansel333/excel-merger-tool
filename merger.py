#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 合并功能模块 - 不使用 pandas，纯 openpyxl 实现
"""

import os
import glob
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter


class ExcelMerger:
    """Excel 文件合并器 - 不依赖 pandas"""
    
    @staticmethod
    def scan_excel_files(directory="."):
        """扫描目录下的所有 Excel 文件"""
        xlsx_files = glob.glob(os.path.join(directory, "*.xlsx"))
        xls_files = glob.glob(os.path.join(directory, "*.xls"))
        
        all_files = xlsx_files + xls_files
        
        # 过滤掉包含"合并结果"的文件
        all_files = [f for f in all_files if "合并结果" not in os.path.basename(f)]
        
        if not all_files:
            return []
        
        # 按文件修改时间排序
        return sorted(all_files, key=lambda x: os.path.getmtime(x))
    
    @staticmethod
    def get_sheet_names(file_path):
        """获取 Excel 文件的 Sheet 名称列表"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            raise Exception(f"无法读取文件 {os.path.basename(file_path)}: {e}")
    
    @staticmethod
    def merge_sheets(files, target_sheet, header_rows_count):
        """
        合并多个 Excel 文件的指定 Sheet
        
        Args:
            files (list): Excel 文件路径列表
            target_sheet (str): 要合并的 Sheet 名称
            header_rows_count (int): 表头行数
            
        Returns:
            tuple: (header_rows, data_rows, file_order) 或 (None, None, None) 如果失败
        """
        header_rows = []
        all_data_rows = []
        file_order = []
        
        for file_path in files:
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                
                if target_sheet not in wb.sheetnames:
                    print(f"  警告: Sheet '{target_sheet}' 在文件 {os.path.basename(file_path)} 中不存在")
                    continue
                
                ws = wb[target_sheet]
                
                # 读取所有行
                all_rows = []
                for row in ws.iter_rows(values_only=True):
                    all_rows.append(list(row) if row else [])
                
                if not all_rows:
                    print(f"  警告: Sheet '{target_sheet}' 为空")
                    wb.close()
                    continue
                
                # 保存表头行
                if not header_rows and len(all_rows) >= header_rows_count:
                    header_rows = all_rows[:header_rows_count]
                    print(f"  保存表头行（前 {header_rows_count} 行）")
                
                # 提取数据行
                if len(all_rows) > header_rows_count:
                    data_rows = all_rows[header_rows_count:]
                    
                    # 过滤掉完全空的行（所有单元格都是None或空字符串）
                    filtered_data_rows = []
                    for row in data_rows:
                        # 检查该行是否有至少一个非空值
                        has_data = False
                        for cell in row:
                            if cell is not None and str(cell).strip() != '':
                                has_data = True
                                break
                        if has_data:
                            filtered_data_rows.append(row)
                    
                    if filtered_data_rows:
                        all_data_rows.extend(filtered_data_rows)
                        file_order.append(os.path.basename(file_path))
                        print(f"  提取了 {len(filtered_data_rows)} 行有效数据（共读取 {len(data_rows)} 行）")
                
                wb.close()
                
            except Exception as e:
                print(f"  错误: 无法读取文件 {os.path.basename(file_path)}: {e}")
                continue
        
        if not all_data_rows or not header_rows:
            return None, None, None
        
        # 重新编号第一列（如果是数字列）
        if all_data_rows:
            first_col_data = [row[0] if row else None for row in all_data_rows if row]
            # 检查第一列是否为数字
            is_numeric = all(
                isinstance(val, (int, float)) or 
                (isinstance(val, str) and val.isdigit())
                for val in first_col_data if val is not None
            )
            
            if is_numeric:
                for i, row in enumerate(all_data_rows, 1):
                    if row:
                        row[0] = i
        
        return header_rows, all_data_rows, file_order
    
    @staticmethod
    def create_output_file(header_rows, data_rows, target_sheet_name, file_order, all_sheet_names, directory="."):
        """
        创建输出 Excel 文件
        
        Args:
            header_rows (list): 表头行数据
            data_rows (list): 数据行
            target_sheet_name (str): 目标 Sheet 名称
            file_order (list): 文件处理顺序
            all_sheet_names (list): 所有 Sheet 名称
            directory (str): 输出目录
            
        Returns:
            str: 输出文件路径，或 None 如果失败
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"合并结果_{target_sheet_name}_{timestamp}.xlsx"
        output_path = os.path.join(directory, output_filename)
        
        try:
            # 创建新工作簿
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # 为每个原有 Sheet 创建空 Sheet
            for sheet_name in all_sheet_names:
                ws = wb.create_sheet(title=sheet_name)
                
                if sheet_name == target_sheet_name:
                    # 写入表头行
                    for header_row in header_rows:
                        ws.append(header_row)
                    
                    # 写入数据行
                    for data_row in data_rows:
                        ws.append(data_row)
                    
                    # 调整列宽
                    for col_idx, column in enumerate(ws.columns, 1):
                        max_length = 0
                        for cell in column:
                            try:
                                cell_len = len(str(cell.value)) if cell.value else 0
                                if cell_len > max_length:
                                    max_length = cell_len
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        col_letter = get_column_letter(col_idx)
                        ws.column_dimensions[col_letter].width = adjusted_width
            
            # 保存文件
            wb.save(output_path)
            wb.close()
            
            print(f"\n合并结果已保存到: {output_filename}")
            print(f"文件结构:")
            end_header_row = len(header_rows)
            if end_header_row == 1:
                print(f"  - 第1行: 表头")
            else:
                print(f"  - 第1-{end_header_row}行: 表头")
            print(f"  - 第{end_header_row + 1}行起: 合并的数据行（共 {len(data_rows)} 行）")
            
            print(f"\n合并时使用的 Excel 文件顺序:")
            for i, filename in enumerate(file_order, 1):
                print(f"  {i}. {filename}")
            
            return output_path
            
        except Exception as e:
            print(f"保存文件时出错: {e}")
            return None
